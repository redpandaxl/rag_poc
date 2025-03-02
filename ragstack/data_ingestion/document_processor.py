"""
Document processing and ingestion for the RAG application.
"""
import os
import shutil
import uuid
from pathlib import Path
from typing import Dict, List, Optional, Any, Union, Callable

from ragstack.config.settings import settings
from ragstack.data_ingestion.text_chunker import TextChunker
from ragstack.models.embeddings import get_embeddings_model
from ragstack.models.metadata import MetadataGenerator
from ragstack.utils.logging import setup_logger
from ragstack.vector_db.factory import get_vector_db


logger = setup_logger("ragstack.data_ingestion.document_processor")


class DocumentProcessor:
    """Process documents and ingest them into the vector database."""
    
    def __init__(
        self,
        chunking_method: str = "recursive",
        vector_db_type: Optional[str] = None
    ):
        """
        Initialize document processor.
        
        Args:
            chunking_method: Method to use for text chunking
            vector_db_type: Type of vector database to use
        """
        self.chunking_method = chunking_method
        self.vector_db_type = vector_db_type or settings.vector_db_type
        
        # Components will be initialized on first use
        self._chunker = None
        self._embedding_model = None
        self._metadata_generator = None
        self._vector_db = None
    
    @property
    def chunker(self) -> Callable:
        """Get the text chunking method."""
        if self._chunker is None:
            self._chunker = TextChunker.get_chunking_method(self.chunking_method)
        return self._chunker
    
    @property
    def embedding_model(self):
        """Get the embedding model."""
        if self._embedding_model is None:
            self._embedding_model = get_embeddings_model()
        return self._embedding_model
    
    @property
    def metadata_generator(self):
        """Get the metadata generator."""
        if self._metadata_generator is None:
            self._metadata_generator = MetadataGenerator()
            self._metadata_generator.initialize()
        return self._metadata_generator
    
    @property
    def vector_db(self):
        """Get the vector database client."""
        if self._vector_db is None:
            self._vector_db = get_vector_db(self.vector_db_type)
        return self._vector_db
    
    def ensure_collection_exists(self, collection_name: str, vector_size: int) -> None:
        """Ensure the collection exists, creating it if it doesn't."""
        if not self.vector_db.collection_exists(collection_name):
            logger.info(f"Creating collection: {collection_name}")
            self.vector_db.create_collection(collection_name, vector_size)
    
    def extract_text_from_file(self, file_path: str) -> str:
        """
        Extract text from a file, handling different file formats.
        
        Args:
            file_path: Path to the file to extract text from
            
        Returns:
            Extracted text content
        
        Raises:
            ValueError: If the file format is unsupported
        """
        file_ext = os.path.splitext(file_path)[1].lower()
        logger.info(f"Extracting text from {file_path} with extension: {file_ext}")
        
        # Text files
        if file_ext in ['.txt', '.md', '.json', '.csv']:
            logger.info(f"Processing as text file: {file_ext}")
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    logger.info(f"Text file read successfully: {len(content)} characters")
                    return content
            except Exception as e:
                logger.error(f"Failed to read text file: {e}", exc_info=True)
                raise ValueError(f"Failed to read text file: {e}")
        
        # Microsoft Word documents
        elif file_ext in ['.docx', '.doc']:
            logger.info(f"Processing as Word document: {file_ext}")
            try:
                logger.info("Attempting to use python-docx")
                import docx
                doc = docx.Document(file_path)
                content = '\n'.join([para.text for para in doc.paragraphs])
                logger.info(f"Word document processed with python-docx: {len(content)} characters")
                return content
            except ImportError:
                logger.warning("python-docx not installed. Trying unstructured...")
                try:
                    logger.info("Attempting to use unstructured")
                    from unstructured.partition.auto import partition
                    elements = partition(file_path)
                    content = "\n".join([str(el) for el in elements])
                    logger.info(f"Word document processed with unstructured: {len(content)} characters")
                    return content
                except ImportError:
                    logger.error("Neither python-docx nor unstructured is installed")
                    raise ValueError("Missing required package: python-docx or unstructured")
                except Exception as e:
                    logger.error(f"Failed to process Word document with unstructured: {e}", exc_info=True)
                    raise ValueError(f"Failed to process Word document: {e}")
            except Exception as e:
                logger.error(f"Failed to process Word document with python-docx: {e}", exc_info=True)
                try:
                    logger.info("Falling back to unstructured after python-docx failure")
                    from unstructured.partition.auto import partition
                    elements = partition(file_path)
                    content = "\n".join([str(el) for el in elements])
                    logger.info(f"Word document processed with unstructured: {len(content)} characters")
                    return content
                except ImportError:
                    logger.error("Unstructured is not installed as fallback")
                    raise ValueError(f"Failed to process Word document: {e}")
                except Exception as e2:
                    logger.error(f"Fallback to unstructured also failed: {e2}", exc_info=True)
                    raise ValueError(f"Failed to process Word document: {e}, fallback also failed: {e2}")
        
        # PDF files
        elif file_ext == '.pdf':
            logger.info("Processing as PDF file")
            try:
                logger.info("Attempting to use PyPDF2")
                import PyPDF2
                with open(file_path, 'rb') as file:
                    reader = PyPDF2.PdfReader(file)
                    logger.info(f"PDF has {len(reader.pages)} pages")
                    text = ""
                    for page_num in range(len(reader.pages)):
                        logger.debug(f"Processing PDF page {page_num+1}/{len(reader.pages)}")
                        page_text = reader.pages[page_num].extract_text()
                        text += page_text + "\n"
                    logger.info(f"PDF processed with PyPDF2: {len(text)} characters")
                    return text
            except ImportError:
                logger.warning("PyPDF2 not installed. Trying unstructured...")
                try:
                    logger.info("Attempting to use unstructured")
                    from unstructured.partition.auto import partition
                    elements = partition(file_path)
                    content = "\n".join([str(el) for el in elements])
                    logger.info(f"PDF processed with unstructured: {len(content)} characters")
                    return content
                except ImportError:
                    logger.error("Neither PyPDF2 nor unstructured is installed")
                    raise ValueError("Missing required package: PyPDF2 or unstructured")
                except Exception as e:
                    logger.error(f"Failed to process PDF with unstructured: {e}", exc_info=True)
                    raise ValueError(f"Failed to process PDF: {e}")
            except Exception as e:
                logger.error(f"Failed to process PDF with PyPDF2: {e}", exc_info=True)
                try:
                    logger.info("Falling back to unstructured after PyPDF2 failure")
                    from unstructured.partition.auto import partition
                    elements = partition(file_path)
                    content = "\n".join([str(el) for el in elements])
                    logger.info(f"PDF processed with unstructured: {len(content)} characters")
                    return content
                except ImportError:
                    logger.error("Unstructured is not installed as fallback")
                    raise ValueError(f"Failed to process PDF: {e}")
                except Exception as e2:
                    logger.error(f"Fallback to unstructured also failed: {e2}", exc_info=True)
                    raise ValueError(f"Failed to process PDF: {e}, fallback also failed: {e2}")
        
        # Excel files
        elif file_ext in ['.xlsx', '.xls']:
            logger.info(f"Processing as Excel file: {file_ext}")
            try:
                logger.info("Attempting to use openpyxl")
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                logger.info(f"Excel file has {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")
                text = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    logger.debug(f"Processing sheet: {sheet_name}")
                    row_count = 0
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " ".join([str(cell) if cell is not None else "" for cell in row])
                        if row_text.strip():  # Only add non-empty rows
                            text.append(row_text)
                            row_count += 1
                    logger.debug(f"Processed {row_count} rows from sheet {sheet_name}")
                content = "\n".join(text)
                logger.info(f"Excel file processed with openpyxl: {len(content)} characters")
                return content
            except ImportError:
                logger.warning("openpyxl not installed. Trying unstructured...")
                try:
                    logger.info("Attempting to use unstructured")
                    from unstructured.partition.auto import partition
                    elements = partition(file_path)
                    content = "\n".join([str(el) for el in elements])
                    logger.info(f"Excel file processed with unstructured: {len(content)} characters")
                    return content
                except ImportError:
                    logger.error("Neither openpyxl nor unstructured is installed")
                    raise ValueError("Missing required package: openpyxl or unstructured")
                except Exception as e:
                    logger.error(f"Failed to process Excel file with unstructured: {e}", exc_info=True)
                    raise ValueError(f"Failed to process Excel file: {e}")
            except Exception as e:
                logger.error(f"Failed to process Excel file with openpyxl: {e}", exc_info=True)
                try:
                    logger.info("Falling back to unstructured after openpyxl failure")
                    from unstructured.partition.auto import partition
                    elements = partition(file_path)
                    content = "\n".join([str(el) for el in elements])
                    logger.info(f"Excel file processed with unstructured: {len(content)} characters")
                    return content
                except ImportError:
                    logger.error("Unstructured is not installed as fallback")
                    raise ValueError(f"Failed to process Excel file: {e}")
                except Exception as e2:
                    logger.error(f"Fallback to unstructured also failed: {e2}", exc_info=True)
                    raise ValueError(f"Failed to process Excel file: {e}, fallback also failed: {e2}")
        
        # Use unstructured as a fallback for other file types
        else:
            logger.info(f"Processing unknown file type: {file_ext}, attempting with unstructured")
            try:
                from unstructured.partition.auto import partition
                elements = partition(file_path)
                content = "\n".join([str(el) for el in elements])
                logger.info(f"File processed with unstructured: {len(content)} characters")
                return content
            except ImportError:
                logger.error("Unstructured is not installed and the file format is not directly supported")
                raise ValueError(f"Unsupported file format: {file_ext} and unstructured not installed")
            except Exception as e:
                logger.error(f"Failed to extract text with unstructured: {e}", exc_info=True)
                raise ValueError(f"Failed to process file: {e}")

    def process_file(self, file_path: str) -> bool:
        """
        Process a single file and ingest into the vector database.
        
        Args:
            file_path: Path to the file to process
            
        Returns:
            True if processing succeeded, False otherwise
        """
        try:
            file_path = os.path.abspath(file_path)
            logger.info(f"START PROCESSING: {file_path}")
            
            # Check file existence and size
            if not os.path.exists(file_path):
                logger.error(f"File not found: {file_path}")
                return False
                
            file_size = os.path.getsize(file_path)
            logger.info(f"File size: {file_size} bytes, type: {os.path.splitext(file_path)[1]}")
            
            # Extract text content based on file type
            logger.info(f"Extracting text from: {file_path}")
            try:
                content = self.extract_text_from_file(file_path)
                logger.info(f"Extraction successful - Content length: {len(content)} characters")
            except Exception as e:
                logger.error(f"Text extraction failed: {e}", exc_info=True)
                raise
            
            if not content.strip():
                logger.warning(f"Extracted empty content from {file_path}")
                return False
            
            # Process and ingest the content
            logger.info(f"Processing content from: {file_path}")
            success = self.process_content(content, file_path)
            logger.info(f"Content processing {'succeeded' if success else 'failed'}")
            
            # If processing succeeded, move the file to processed directory
            if success:
                processed_path = os.path.join(
                    settings.processed_data_dir, 
                    os.path.basename(file_path)
                )
                logger.info(f"Moving file to processed directory: {processed_path}")
                try:
                    # Check if file still exists (to handle race conditions)
                    if os.path.exists(file_path):
                        # Use copy then delete to avoid cross-device issues
                        shutil.copy2(file_path, processed_path)
                        os.remove(file_path)
                        logger.info(f"Successfully moved processed file to: {processed_path}")
                    else:
                        logger.warning(f"File no longer exists at {file_path}, may have been moved by another process")
                except Exception as move_error:
                    logger.warning(f"Error moving processed file: {move_error}")
                    # Still consider it a success if content was processed
            
            logger.info(f"FINISHED PROCESSING: {file_path} - Result: {'SUCCESS' if success else 'FAILURE'}")
            return success
        
        except Exception as e:
            logger.error(f"Failed to process file {file_path}: {e}", exc_info=True)
            
            # Move to failed directory if the file exists
            try:
                if os.path.exists(file_path):
                    failed_path = os.path.join(
                        settings.failed_data_dir, 
                        os.path.basename(file_path)
                    )
                    logger.info(f"Moving file to failed directory: {failed_path}")
                    # Use copy then delete to avoid cross-device issues
                    shutil.copy2(file_path, failed_path)
                    os.remove(file_path)
                    logger.info(f"Moved failed file to: {failed_path}")
                else:
                    logger.warning(f"File no longer exists at {file_path}, may have been moved by another process")
            except Exception as move_error:
                logger.warning(f"Error moving failed file: {move_error}")
            
            logger.info(f"FINISHED PROCESSING: {file_path} - Result: EXCEPTION")
            return False
    
    def process_content(self, content: str, source_info: str = "") -> bool:
        """
        Process content and ingest into the vector database.
        
        Args:
            content: Text content to process
            source_info: Source information for logging
            
        Returns:
            True if processing succeeded, False otherwise
        """
        # Generate unique document ID for metrics tracking
        document_id = str(uuid.uuid4())
        document_name = os.path.basename(source_info) if os.path.exists(source_info) else "content"
        document_type = os.path.splitext(source_info)[1].lstrip('.') if os.path.exists(source_info) else "text"
        
        # Import metrics components if available
        try:
            from ragstack.utils.monitoring.metrics import metrics_collector, ProcessingStage
            has_metrics = True
            # Start metrics tracking
            metrics_collector.start_document_processing(document_id, document_name, document_type)
        except ImportError:
            has_metrics = False
            logger.debug("Metrics collection not available")
        
        try:
            logger.info(f"Processing content from: {source_info}")
            
            # Check for services that we need
            try:
                logger.info("Checking embedding model availability...")
                is_embedding_available = self.embedding_model is not None
                logger.info(f"Embedding model available: {is_embedding_available}")
                
                logger.info("Checking vector database availability...")
                vector_db_name = self.vector_db_type
                is_vector_db_available = self.vector_db is not None
                logger.info(f"Vector database ({vector_db_name}) available: {is_vector_db_available}")
            except Exception as e:
                logger.error(f"Service availability check failed: {e}", exc_info=True)
            
            # Chunk the content
            logger.info(f"Chunking content with method: {self.chunking_method}")
            start_time = __import__('time').time()
            chunks = self.chunker(content)
            chunking_time = __import__('time').time() - start_time
            logger.info(f"Created {len(chunks)} chunks in {chunking_time:.2f} seconds from source: {source_info}")
            
            # Record chunking metrics if available
            if has_metrics:
                metrics_collector.record_stage_timing(
                    document_id, 
                    ProcessingStage.CHUNKING, 
                    chunking_time,
                    success=True,
                    metadata={
                        "chunks_count": len(chunks),
                        "chunking_method": self.chunking_method
                    }
                )
            
            # Skip empty content
            if not chunks:
                logger.warning(f"No chunks created from source: {source_info}")
                if has_metrics:
                    metrics_collector.complete_document_processing(
                        document_id, False, 0, "No chunks created"
                    )
                return False
            
            # Log chunk statistics
            chunk_sizes = [len(chunk) for chunk in chunks]
            avg_chunk_size = sum(chunk_sizes) / len(chunk_sizes)
            min_chunk_size = min(chunk_sizes)
            max_chunk_size = max(chunk_sizes)
            logger.info(f"Chunk statistics - Count: {len(chunks)}, Avg size: {avg_chunk_size:.1f}, Min: {min_chunk_size}, Max: {max_chunk_size}")
            
            # Process chunks
            documents = []
            vectors = []
            metadatas = []
            
            logger.info("Processing chunks...")
            metadata_time_total = 0
            embedding_time_total = 0
            
            for i, chunk in enumerate(chunks):
                logger.debug(f"Processing chunk {i+1}/{len(chunks)} from {source_info}")
                
                # Generate metadata
                try:
                    logger.debug(f"Generating metadata for chunk {i+1}")
                    start_time = __import__('time').time()
                    metadata = self.metadata_generator.generate(chunk)
                    metadata_time = __import__('time').time() - start_time
                    metadata_time_total += metadata_time
                    logger.debug(f"Metadata generated in {metadata_time:.2f} seconds")
                except Exception as e:
                    logger.error(f"Failed to generate metadata for chunk {i+1}: {e}", exc_info=True)
                    metadata = {"error": str(e)}  # Create basic metadata to continue
                
                # Add source and chunk information to metadata
                metadata["source"] = source_info
                metadata["chunk_index"] = i
                metadata["chunk_count"] = len(chunks)
                metadata["document_id"] = document_id
                
                # Generate embedding
                try:
                    logger.debug(f"Generating embedding for chunk {i+1}")
                    start_time = __import__('time').time()
                    embedding = self.embedding_model.encode_query(chunk)
                    embedding_time = __import__('time').time() - start_time
                    embedding_time_total += embedding_time
                    logger.debug(f"Embedding generated in {embedding_time:.2f} seconds, dimensions: {len(embedding)}")
                except Exception as e:
                    logger.error(f"Failed to generate embedding for chunk {i+1}: {e}", exc_info=True)
                    if has_metrics:
                        metrics_collector.record_stage_timing(
                            document_id, 
                            ProcessingStage.EMBEDDING, 
                            0,
                            success=False,
                            error_message=str(e)
                        )
                        metrics_collector.complete_document_processing(
                            document_id, False, i, f"Embedding generation failed: {e}"
                        )
                    raise  # This is critical, so we should fail
                
                # Add to lists
                documents.append(chunk)
                vectors.append(embedding)
                metadatas.append(metadata)
                
                if (i+1) % 10 == 0 or i+1 == len(chunks):
                    logger.info(f"Processed {i+1}/{len(chunks)} chunks")
            
            # Record metadata and embedding metrics if available
            if has_metrics:
                metrics_collector.record_stage_timing(
                    document_id, 
                    ProcessingStage.METADATA, 
                    metadata_time_total,
                    success=True
                )
                metrics_collector.record_stage_timing(
                    document_id, 
                    ProcessingStage.EMBEDDING, 
                    embedding_time_total,
                    success=True,
                    metadata={"embedding_dimensions": len(vectors[0]) if vectors else 0}
                )
            
            # Ensure collection exists
            collection_name = settings.vector_db_settings[self.vector_db_type]["collection_name"]
            vector_size = settings.embedding_model["vector_size"]
            logger.info(f"Ensuring vector collection exists: {collection_name} with size {vector_size}")
            self.ensure_collection_exists(collection_name, vector_size)
            
            # Insert into vector database
            logger.info(f"Inserting {len(chunks)} chunks into vector database")
            start_time = __import__('time').time()
            self.vector_db.insert(
                collection_name=collection_name,
                vectors=vectors,
                documents=documents,
                metadata=metadatas
            )
            insert_time = __import__('time').time() - start_time
            logger.info(f"Successfully ingested {len(chunks)} chunks into {collection_name} in {insert_time:.2f} seconds")
            
            # Record vector DB metrics if available
            if has_metrics:
                metrics_collector.record_stage_timing(
                    document_id, 
                    ProcessingStage.VECTORDB, 
                    insert_time,
                    success=True,
                    metadata={"collection_name": collection_name}
                )
                metrics_collector.complete_document_processing(
                    document_id, True, len(chunks)
                )
                
            return True
        
        except Exception as e:
            logger.error(f"Failed to process content from {source_info}: {e}", exc_info=True)
            if has_metrics:
                metrics_collector.complete_document_processing(
                    document_id, False, 0, str(e)
                )
            return False